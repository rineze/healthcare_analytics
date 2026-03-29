"""Export Medicaid managed care plans to a clean, formatted Excel file."""

import sys
sys.path.insert(0, r"C:\dev\healthcare_data")

from db import get_conn, LOCAL
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).parent / "Medicaid_MCO_Reference_Table.xlsx"

QUERY = """
SELECT
    state AS "State",
    plan_name AS "Plan Name",
    plan_year AS "Plan Year",
    plan_type AS "Plan Type",
    benefit_category AS "Benefit Category",
    program_type AS "Program Type",
    parent_organization AS "Parent Organization",
    program_name AS "CMS Program Name",
    geographic_region AS "Geographic Region",
    medicaid_enrollment AS "Medicaid Enrollment",
    dual_enrollment AS "Dual Enrollment",
    total_enrollment AS "Total Enrollment"
FROM drinf.ref_medicaid_landscape
ORDER BY state, benefit_category, plan_name
"""


def build():
    conn = get_conn(LOCAL)
    df = pd.read_sql(QUERY, conn)
    conn.close()

    # Write to Excel with openpyxl for formatting
    with pd.ExcelWriter(str(OUTPUT), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Medicaid Plans")
        ws = writer.sheets["Medicaid Plans"]

        # --- Styling ---
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        body_font = Font(name="Calibri", size=10)
        body_align = Alignment(vertical="top", wrap_text=True)
        number_align = Alignment(horizontal="right", vertical="top")

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        # Style header row
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Style data rows
        num_cols = {"Medicaid Enrollment", "Dual Enrollment", "Total Enrollment", "Plan Year"}
        for row_idx in range(2, len(df) + 2):
            row_fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.border = thin_border
                col_name = df.columns[col_idx - 1]
                if col_name in num_cols:
                    cell.alignment = number_align
                    if col_name != "Plan Year":
                        cell.number_format = "#,##0"
                else:
                    cell.alignment = body_align
                if row_fill:
                    cell.fill = row_fill

        # Auto-size columns (with sensible max widths)
        max_widths = {
            "State": 8,
            "Plan Name": 45,
            "Plan Year": 11,
            "Plan Type": 12,
            "Benefit Category": 20,
            "Program Type": 15,
            "Parent Organization": 35,
            "CMS Program Name": 40,
            "Geographic Region": 45,
            "Medicaid Enrollment": 20,
            "Dual Enrollment": 17,
            "Total Enrollment": 17,
        }
        for col_idx, col_name in enumerate(df.columns, 1):
            # Calculate width from data (header + longest value)
            max_len = len(col_name)
            for val in df[col_name].astype(str).head(100):
                max_len = max(max_len, len(str(val)))
            # Cap at defined max
            cap = max_widths.get(col_name, 30)
            width = min(max_len + 2, cap)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze top row + add auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Set row height for header
        ws.row_dimensions[1].height = 30

    print(f"Saved {len(df)} plans to {OUTPUT}")


if __name__ == "__main__":
    build()
